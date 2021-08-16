import openpyxl
from openpyxl.styles import Border, Side, Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


class Excel:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.worksheets[0]

    def add_border(self, cell_range):
        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

        rows = [rows for rows in self.ws[cell_range]]
        flattened = [item for sublist in rows for item in sublist]
        [(setattr(cell, 'border', border)) for cell in
         flattened]

    def highlight_row(self, rows, columns, color):
        for column in range(1, columns):
            self.ws.cell(row=rows, column=column).fill = PatternFill(start_color=color,
                                                                     end_color=color,
                                                                     fill_type='solid')

    def width_as_header(self):
        for column_cells in self.ws.columns:
            length = len(str(column_cells[0].value))
            self.ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    def make_bold(self, row, column):
        for row in range(1, row + 1):
            self.ws[get_column_letter(row) + str(column)].font = Font(bold=True)
