import os
import re
import smtplib
import time
from email import encoders
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from platform import python_version

import openpyxl
from openpyxl.styles import Border, Side, Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from selenium import webdriver


class Excel:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.worksheets[0]

    def add_border(self, cell_range):
        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

        rows = [rows for rows in self.ws[cell_range]]
        flattened = [item for sublist in rows for item in sublist]
        [(setattr(cell, 'border', border)) for cell in
         flattened]

    def highlight_row(self, rows, columns, color):
        for column in range(1, columns):
            self.ws.cell(row=rows, column=column).fill = PatternFill(start_color=color,
                                                                     end_color=color,
                                                                     fill_type='solid')

    def width_as_header(self):
        for column_cells in self.ws.columns:
            length = len(str(column_cells[0].value))
            self.ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    def make_bold(self, row, column):
        for row in range(1, row + 1):
            self.ws[get_column_letter(row) + str(column)].font = Font(bold=True)


class Email:
    def __init__(self, server, user, password, recipients, subject, text):
        self.server = server
        self.user = user
        self.password = password
        self.recipients = recipients
        self.subject = subject
        self.text = text
        self.msg = MIMEMultipart('alternative')

    def generate_mail(self):
        self.msg['Subject'] = self.subject
        self.msg['From'] = 'Python script <' + self.user + '>'
        self.msg['To'] = ', '.join(self.recipients)
        self.msg['Reply-To'] = self.user
        self.msg['Return-Path'] = self.user
        self.msg['X-Mailer'] = 'Python/' + (python_version())
        part_text = MIMEText(self.text, 'plain')
        self.msg.attach(part_text)

    def add_file(self, filepath):
        basename = os.path.basename(filepath)
        filesize = os.path.getsize(filepath)
        part_file = MIMEApplication(open(filepath, "rb").read(), Name=os.path.basename(filepath), _subtype="xlsl")
        part_file.set_payload(open(filepath, "rb").read())
        part_file.add_header('Content-Description', basename)
        part_file.add_header('Content-Disposition', 'attachment; filename="{}"; size={}'.format(basename, filesize))
        encoders.encode_base64(part_file)
        self.msg.attach(part_file)

    def add_html(self):
        html = '<html><head></head><body><p>' + self.text + '</p></body></html>'
        part_html = MIMEText(html, 'html')
        self.msg.attach(part_html)

    def send_mail(self):
        mail = smtplib.SMTP_SSL(self.server)
        mail.login(self.user, self.password)
        mail.sendmail(self.user, self.recipients, self.msg.as_string())
        mail.quit()


def get_end(l):
    l = l % 100
    if (l > 20):
        return get_end(l % 10)

    if (l == 1):
        return 'а'
    elif (2 <= l <= 4):
        return 'и'
    else:
        return ''


def get_driver():
    chrome_options = webdriver.ChromeOptions()
    # chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--log-level=3")
    chrome_options.add_argument("--start-maximized")
    return webdriver.Chrome(options=chrome_options)


class Ozon:
    def parse_page(self, driver):
        blocks = driver.find_elements_by_xpath("//div[@class='e3f5']")
        for i in range(8):
            driver.execute_script(f"scroll ({1080 * i}, {1080 * i + 1080});")
            time.sleep(0.5)

        data = []
        for block in blocks:
            name = block.find_element_by_xpath(
                ".//span[contains(@class,'j4 as3 az a0f2 f-tsBodyL item e3t0')]/span").get_attribute('innerHTML')
            price_block = block.find_elements_by_xpath(".//div[contains(@class,'b5v4 e3r9 item')]/span")
            if len(price_block) == 1:
                full_price = discount_price = price_block[0].get_attribute('innerHTML')
            else:
                full_price = price_block[1].get_attribute('innerHTML')
                discount_price = price_block[0].get_attribute('innerHTML')

            try:
                review = int(
                    re.findall('\d+', block.find_element_by_xpath(".//a[@class='a1r7']").get_attribute('innerHTML'))[0])
            except:
                review = '0'

            try:
                description = block.find_element_by_xpath(".//span[@class='j4 as3 a0f6 f-tsBodyM item e3t']/span").text
            except:
                description = ''
            data.append([name, description, review, re.sub(r'\s+', '', discount_price), re.sub(r'\s+', '', full_price)])
        return data

    def parse_query(self, user_query):
        driver = get_driver()
        try:
            driver.get("https://www.ozon.ru/")
            driver.find_element_by_xpath("//div[@class='f9j4']/input").send_keys(user_query)
            time.sleep(1)
            driver.find_element_by_xpath("//button[@class='f9k']").click()
            time.sleep(3)
            try:
                total = re.sub('[^0-9]', '', driver.find_element_by_class_name('b6r7').text)
                url = driver.current_url
                driver.get(url[:url.find('&') + 1] + 'sorting=rating' + '&' + url[url.find('&') + 1:])
                data = []
                for i in range(3):
                    time.sleep(3)
                    data.append(self.parse_page(driver))
                    try:
                        driver.find_element_by_class_name('b7t1').find_element_by_class_name('kxa6').click()
                    except:
                        break
                return data, total
            except Exception as e:
                print(e)
                print("По данному запросу ничего не найдено!")
                return None, None
        except Exception as e:
            print(e)
            print("Произошла ошибка попробуйте повторить")
            return None, None
        finally:
            driver.close()


if __name__ == '__main__':
    while True:
        user_query = str(input("Введите запрос: "))
        data, total = Ozon().parse_query(user_query)
        if data is not None:
            break

    data = [x for sublist in data for x in sublist]
    excel = Excel()
    excel.ws.append(("Наименование", "Описание", "Количество отзывов", "Цена со скидкой", "Цена без скидки"))
    for i in data:
        excel.ws.append(i)
    excel.ws.append([f'Общее количество товаров: {total}'])

    excel.ws.merge_cells(f'A{len(data) + 2}:E{len(data) + 2}')

    price = [int(re.sub('[^0-9]', '', i[4])) for i in data]
    all_mins = [index for index, value in enumerate(price) if value == min(price)]
    for row in all_mins:
        excel.highlight_row(row + 2, len(data[0]) + 1, 'FF0000')

    excel.width_as_header()
    excel.add_border(cell_range=f'A1:E{len(data) + 1}')

    excel.make_bold(len(data[0]), 1)
    excel.make_bold(len(data[0]), len(data) + 2)
    excel.wb.save("Шаблон для заполнения.xlsx")
    email = Email('smtp.gmail.com', 'email@gmail.com', "password", ['email2@gmail.com'], 'Отчет',
                  f"Запрос пользователя: {user_query}\nВ файле {len(data)} строк" + get_end(len(data))) # заменить email@gmail.com на почту отправителя,
    # заменить password на пароль от почты отправитлея, заменить ['email2@gmail.com'] на список почт получателей
    email.generate_mail()
    email.add_file("Шаблон для заполнения.xlsx")
    email.send_mail()
